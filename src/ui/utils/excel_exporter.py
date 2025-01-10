import openpyxl
from openpyxl.styles import PatternFill

class ExcelExporter:
    def export_to_excel(self, worker_data, output_path):
        """ワーカーのスキルデータをExcelファイルに出力する。"""
        print("--- ExcelExporter: export_to_excel ---")
        print(f"  worker_data: {worker_data}")

        all_categories = []
        for worker_name, worker_obj in worker_data.items():
            for category in worker_obj.categories:
                if category not in all_categories:
                    all_categories.append(category)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Skill Matrix"

        # ヘッダー行を作成
        headers = ["名前"] + all_categories
        ws.append(headers)

        # データ行を作成
        for worker_name, worker_obj in worker_data.items():
            row = [worker_name]
            for category in all_categories:
                if category in worker_obj.categories:
                    index = worker_obj.categories.index(category)
                    row.append(worker_obj.skill_levels[index])
                else:
                    row.append("")  # カテゴリーがない場合は空欄
            ws.append(row)

        # スタイルの設定（ここでは、スキルレベルに応じてセルの色を変更）
        for row in ws.iter_rows(min_row=2):  # ヘッダー行を除外
            for cell in row:
                if isinstance(cell.value, int):
                    if cell.value == 5:
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
                    elif cell.value == 4:
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                    elif cell.value == 3:
                        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
                    elif cell.value == 2:
                        cell.fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Pink
                    else:
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
        print(f"Saving Excel file to: {output_path}")
        wb.save(output_path)